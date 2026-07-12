#!/usr/bin/env python3
"""Recalculate every formula in an .xlsx via headless LibreOffice, then scan for
Excel error cells and print a JSON report.

Libraries write formula STRINGS but never compute them, so a freshly written
workbook has empty cached values (openpyxl data_only=True reads None). This opens
the file in LibreOffice, runs a Basic macro that calls calculateAll() and saves,
then reloads with openpyxl to (a) count formulas and (b) find any #REF!/#DIV/0!/
etc. cells.

Usage:  recalc.py <file.xlsx> [timeout_seconds]   (default timeout 60)

Output JSON:
  {"status": "success"|"errors_found", "total_formulas": N, "total_errors": N,
   "error_summary": {"#REF!": {"count": 2, "locations": ["Sheet1!B5", ...]}}}
On failure: {"status": "error", "message": "..."}  (exit 1)
"""
from __future__ import annotations

import json
import os
import platform
import shutil
import subprocess
import sys
from pathlib import Path

ERROR_STRINGS = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!")

MACRO = (
    "Sub RecalculateAndSave\n"
    "    ThisComponent.calculateAll()\n"
    "    ThisComponent.store()\n"
    "    ThisComponent.close(True)\n"
    "End Sub\n"
)

MODULE_XBA = (
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" '
    '"module.dtd">\n'
    '<script:module xmlns:script="http://openoffice.org/2000/script" '
    'script:name="Module1" script:language="StarBasic">'
    + MACRO.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    + "</script:module>\n"
)


def _fail(message: str) -> int:
    print(json.dumps({"status": "error", "message": message}))
    return 1


def _macro_dir() -> Path:
    if platform.system() == "Darwin":
        base = Path.home() / "Library/Application Support/LibreOffice/4/user"
    else:
        base = Path.home() / ".config/libreoffice/4/user"
    return base / "basic/Standard"


def _install_macro(soffice: str) -> None:
    """Ensure the RecalculateAndSave macro exists in the user profile."""
    md = _macro_dir()
    if not md.parent.parent.exists():
        # No profile yet: let LibreOffice create one, then install the macro.
        subprocess.run(
            [soffice, "--headless", "--terminate_after_init"],
            capture_output=True, timeout=120,
        )
    md.mkdir(parents=True, exist_ok=True)
    (md / "Module1.xba").write_text(MODULE_XBA)
    script_lb = md / "script.xlb"
    if not script_lb.exists():
        script_lb.write_text(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument '
            '1.0//EN" "library.dtd">\n'
            '<library:library xmlns:library="http://openoffice.org/2000/library" '
            'library:name="Standard" library:readonly="false" '
            'library:passwordprotected="false">\n'
            ' <library:element library:name="Module1"/>\n'
            '</library:library>\n'
        )


def _recalc(soffice: str, path: Path, timeout: int) -> None:
    macro_url = (
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave"
        "?language=Basic&location=application"
    )
    cmd = [soffice, "--headless", "--norestore", macro_url, str(path)]
    gtimeout = shutil.which("gtimeout") or shutil.which("timeout")
    if gtimeout:
        cmd = [gtimeout, str(timeout)] + cmd
    proc = subprocess.run(cmd, capture_output=True, timeout=timeout + 30)
    # 124 = wrapper timeout (LibreOffice often finishes the store anyway); tolerate.
    if proc.returncode not in (0, 124):
        raise RuntimeError(
            f"soffice exited {proc.returncode}: {proc.stderr.decode(errors='replace')[:400]}"
        )


def _scan(path: Path) -> dict:
    from openpyxl import load_workbook

    summary: dict[str, dict] = {}
    total_errors = 0
    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v in ERROR_STRINGS:
                    total_errors += 1
                    e = summary.setdefault(v, {"count": 0, "locations": []})
                    e["count"] += 1
                    if len(e["locations"]) < 20:
                        e["locations"].append(f"{ws.title}!{cell.coordinate}")
    wb.close()

    total_formulas = 0
    wf = load_workbook(path)
    for ws in wf.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total_formulas += 1
    wf.close()

    return {
        "status": "errors_found" if total_errors else "success",
        "total_formulas": total_formulas,
        "total_errors": total_errors,
        "error_summary": summary,
    }


def main() -> int:
    if len(sys.argv) < 2:
        return _fail("usage: recalc.py <file.xlsx> [timeout_seconds]")
    path = Path(sys.argv[1]).resolve()
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 60
    if not path.is_file():
        return _fail(f"no such file: {path}")

    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return _fail(
            "LibreOffice not found. Install it: brew install --cask libreoffice "
            "(macOS) or apt-get install libreoffice-calc (Linux)."
        )
    try:
        from openpyxl import load_workbook  # noqa: F401
    except ImportError:
        return _fail("openpyxl not installed. Run: pip install openpyxl")

    try:
        _install_macro(soffice)
        _recalc(soffice, path, timeout)
    except subprocess.TimeoutExpired:
        return _fail(f"LibreOffice timed out after {timeout}s (raise the timeout arg)")
    except Exception as e:  # noqa: BLE001 - surface any driver failure cleanly
        return _fail(str(e))

    report = _scan(path)
    print(json.dumps(report, indent=2))
    return 0


if __name__ == "__main__":
    # LibreOffice needs a writable HOME for its profile.
    os.environ.setdefault("HOME", str(Path.home()))
    sys.exit(main())
