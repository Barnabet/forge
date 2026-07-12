#!/usr/bin/env python3
"""Summarize an .xlsx: per-sheet dimensions, headers, a few sample rows, and any
Excel error cells. Uses cached values (data_only), so run recalc.py first if the
workbook was just written by a library.

Usage: xlsx_inspect.py <file.xlsx> [sample_rows]   (default 5 sample rows)
"""
from __future__ import annotations

import sys
from pathlib import Path

ERROR_STRINGS = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}


def _cell(v) -> str:
    if v is None:
        return ""
    s = str(v)
    return s if len(s) <= 24 else s[:21] + "..."


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: xlsx_inspect.py <file.xlsx> [sample_rows]", file=sys.stderr)
        return 1
    path = Path(sys.argv[1])
    sample = int(sys.argv[2]) if len(sys.argv) > 2 else 5
    if not path.is_file():
        print(f"error: no such file: {path}", file=sys.stderr)
        return 1

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        return 1

    wb = load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        print(f"== {ws.title}  ({ws.max_row} rows x {ws.max_column} cols) ==")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("  (empty)\n")
            continue
        print("  headers: " + " | ".join(_cell(v) for v in rows[0]))
        for r in rows[1 : 1 + sample]:
            print("  row:     " + " | ".join(_cell(v) for v in r))
        if len(rows) > 1 + sample:
            print(f"  ... {len(rows) - 1 - sample} more data rows")

        errors = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in ERROR_STRINGS:
                    errors.append(f"{cell.coordinate}={cell.value}")
        if errors:
            print("  ERROR CELLS: " + ", ".join(errors[:20])
                  + (f" (+{len(errors) - 20} more)" if len(errors) > 20 else ""))
        print()
    wb.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
